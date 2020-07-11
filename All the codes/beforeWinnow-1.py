import xlrd
import xlwt
import openpyxl
from openpyxl import load_workbook

"""
loc = ("C:/UMD/Research/twitter eligible codes.xlsx")
myWb = xlrd.open_workbook(loc)
sheet = myWb.sheet_by_index(0)
sheet.cell_value(0, 0)
tweets = []
for i in range(sheet.nrows):
    tweets.append(sheet.cell_value(i, 0))
"""
n=1
wb = openpyxl.load_workbook('C:/UMD/Research/Before Winnow-2.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
for myCol in range(2, 1000001):
    ws.cell(row=myCol, column= 1).value = n
    n = n+1
wb.save('C:/UMD/Research/Before Winnow-2.xlsx')
