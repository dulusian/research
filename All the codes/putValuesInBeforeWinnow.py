import pymysql
import xlrd
import xlwt
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from pymysql import IntegrityError

myDB = pymysql.connect(host="127.0.0.1", user="root", passwd="f9d0s-jooN", db="tweetdb")
cur = myDB.cursor()
putWordToDB = "insert into tweetsWithTopicsAndCount (tweet_id, gram_id, count, topicNum) values (%s, %s, %s, %s);"


loc = ("C:/UMD/Research/topicsWithTweetID.xlsx")
myWb = xlrd.open_workbook(loc)
sheet = myWb.sheet_by_index(0)
sheet.cell_value(0, 0)
tweets = []
topics = []
for i in range(1,sheet.nrows):
    if sheet.cell_value(i,0) not in tweets:
        tweets.append(int(sheet.cell_value(i, 0)))
    #print int(sheet.cell_value(i,0))
        topics.append(int(sheet.cell_value(i, 1)))
    #print int(sheet.cell_value(i,1))

loc = ("C:/UMD/Research/CountGramTweetResults/CountGram.xlsx")
myWb = xlrd.open_workbook(loc)
sheetOfCountGram = myWb.sheet_by_index(0)
sheetOfCountGram.cell_value(0, 0)
count = []
gramID = []
tweetID = []
for i in range(1, sheetOfCountGram.nrows):
    count.append(int(sheetOfCountGram.cell_value(i, 2)))
    gramID.append(int(sheetOfCountGram.cell_value(i, 1)))
    tweetID.append(int(sheetOfCountGram.cell_value(i, 0)))
    #print '[TWEETID, GRAM ID, COUNT] =', int(sheetOfCountGram.cell_value(i, 0)), int(sheetOfCountGram.cell_value(i, 1)), int(sheetOfCountGram.cell_value(i, 2))
#wb = openpyxl.load_workbook('C:/UMD/Research/EligibleTweetsWithCount.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')

for i in range(0, len(tweets)):
    for j in range(0, len(tweetID)):
        if tweets[i] == tweetID[j]:
            #print 'Final -', tweets[i], gramID[j], count[j], topics[i]
            #print ''
            try:
                cur.execute(putWordToDB, [tweets[i], gramID[j], count[j], topics[i]])
            except IntegrityError:
                print 'Integrity Error on ', tweets[i],'th index'


myDB.commit()

            #ws.cell(row = gramID[myRow] +1, column = myCol+1).value = count[myRow]
                #print "cell Row : ", gramID[myRow]+1, " column : ", myCol
                #print "gramID", gramID[myRow] + 1
                #print "count", count[myRow]
#wb.save('C:/UMD/Research/EligibleTweetsWithCount.xlsx')