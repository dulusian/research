import xlrd
import xlwt
import openpyxl
from openpyxl import load_workbook

myCol = 1
wb = openpyxl.load_workbook('C:/UMD/Research/Before Winnow-2.xlsx')
num = 1
for ws in wb.worksheets:
    for myRow in range(2,1000001):
        ws.cell(row=myRow, column=myCol).value = num
        num = num+1

wb.save('C:/UMD/Research/Before Winnow-2.xlsx')
wb.close()
